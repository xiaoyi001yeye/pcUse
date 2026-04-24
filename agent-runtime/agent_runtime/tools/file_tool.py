from __future__ import annotations

import csv
import json
import os
import subprocess
from pathlib import Path
from typing import Any

from agent_runtime.schemas.tool_call import ToolCall, ToolResult


class FileTool:
    name = "file"

    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir

    def run(self, call: ToolCall) -> ToolResult:
        action = call.action
        if action == "open":
            return self.open(call.args.get("path"))
        if action == "read_preview":
            return self.read_preview(call.args.get("path"), int(call.args.get("rows", 5)))
        if action == "list_dir":
            return self.list_dir(call.args.get("path"))
        if action == "exists":
            return self.exists(call.args.get("path"))
        return ToolResult(ok=False, error=f"Unsupported file action: {action}")

    def open(self, path_value: Any) -> ToolResult:
        path = self._path(path_value)
        if not path.exists():
            return ToolResult(ok=False, error=f"File not found: {path}")
        try:
            if os.name == "nt":
                os.startfile(str(path))  # type: ignore[attr-defined]
            elif os.name == "posix":
                subprocess.Popen(["xdg-open", str(path)])
            else:
                return ToolResult(ok=False, error="Open file is not supported on this platform")
            return ToolResult(ok=True, output={"path": str(path), "opened": True})
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=str(exc))

    def read_preview(self, path_value: Any, rows: int = 5) -> ToolResult:
        path = self._path(path_value)
        if not path.exists():
            return ToolResult(ok=False, error=f"File not found: {path}")
        rows = max(1, min(50, rows))
        suffix = path.suffix.lower()
        try:
            if suffix in {".txt", ".md", ".json", ".log", ".py", ".ts", ".tsx"}:
                lines = path.read_text(encoding="utf-8", errors="replace").splitlines()[:rows]
                return ToolResult(ok=True, output={"path": str(path), "rows": lines})
            if suffix == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as fh:
                    reader = csv.reader(fh)
                    sample = [row for _, row in zip(range(rows), reader)]
                return ToolResult(ok=True, output={"path": str(path), "rows": sample})
            if suffix in {".xlsx", ".xlsm"}:
                try:
                    from openpyxl import load_workbook
                except Exception as exc:  # noqa: BLE001
                    return ToolResult(ok=False, error=f"openpyxl is required to read Excel files: {exc}")
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                sample = []
                for row in ws.iter_rows(max_row=rows, values_only=True):
                    sample.append([cell for cell in row])
                return ToolResult(ok=True, output={"path": str(path), "sheet": ws.title, "rows": sample})
            return ToolResult(ok=False, error=f"Preview not supported for suffix: {suffix}")
        except Exception as exc:  # noqa: BLE001
            return ToolResult(ok=False, error=str(exc))

    def list_dir(self, path_value: Any) -> ToolResult:
        path = self._path(path_value)
        if not path.exists() or not path.is_dir():
            return ToolResult(ok=False, error=f"Directory not found: {path}")
        items = [{"name": p.name, "is_dir": p.is_dir(), "size": p.stat().st_size if p.is_file() else None} for p in path.iterdir()]
        return ToolResult(ok=True, output={"path": str(path), "items": items})

    def exists(self, path_value: Any) -> ToolResult:
        path = self._path(path_value)
        return ToolResult(ok=True, output={"path": str(path), "exists": path.exists(), "is_file": path.is_file(), "is_dir": path.is_dir()})

    def _path(self, value: Any) -> Path:
        if not value:
            raise ValueError("path is required")
        return Path(str(value)).expanduser().resolve()
